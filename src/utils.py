"""Utility classes and functions for the dictionary application."""
import json
import re
import logging
from typing import List, Dict, Tuple, Any
from pathlib import Path

import openpyxl
from openpyxl import Workbook, load_workbook

from .models import DictionaryEntry, Translation

logger = logging.getLogger(__name__)

# ------------------ Constants ------------------

STATUS_LABELS = {
    "not_started": "Öğrenilmedi",
    "learning": "Öğreniliyor",
    "learned": "Öğrenildi",
}

STATUS_COLORS = {
    "not_started": "#9e9e9e",
    "learning": "#ff9800",
    "learned": "#4caf50",
}

LANGUAGE_NAMES = {
    "tr": "Türkçe",
    "ru": "Rusça",
    "en": "İngilizce",
}

# ------------------ Settings Manager ------------------

class SettingsManager:
    DEFAULTS = {
        "tts_enabled": True,
        "auto_tts": False,
        "tts_rate": 0,
        "theme": "light",
        "window_geometry": None,
    }

    def __init__(self, filepath: str = "settings.json"):
        self.filepath = Path(filepath)
        self._settings: Dict[str, Any] = {}
        self.load()

    def load(self):
        if self.filepath.exists():
            try:
                with open(self.filepath, "r", encoding="utf-8") as f:
                    loaded = json.load(f)
                    if isinstance(loaded, dict):
                        self._settings = loaded
                    else:
                        self._settings = {}
            except (json.JSONDecodeError, OSError, TypeError, ValueError) as e:
                logger.warning(f"Settings load failed: {e}, using defaults.")
                self._settings = {}
        for key, val in self.DEFAULTS.items():
            if key not in self._settings:
                self._settings[key] = val

    def save(self):
        try:
            with open(self.filepath, "w", encoding="utf-8") as f:
                json.dump(self._settings, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"Settings save error: {e}")

    def get(self, key: str, default=None):
        return self._settings.get(key, default)

    def set(self, key: str, value):
        self._settings[key] = value
        self.save()


# ------------------ Collation / Sorting ------------------

TURKISH_ALPHABET = "abcçdefgğhıijklmnoöprsştuüvyz"
TURKISH_ORDER = {ch: idx for idx, ch in enumerate(TURKISH_ALPHABET)}
RUSSIAN_ALPHABET = "абвгдеёжзийклмнопрстуфхцчшщъыьэюя"
RUSSIAN_ORDER = {ch: idx for idx, ch in enumerate(RUSSIAN_ALPHABET)}


def turkish_collation_key(s: str) -> List[int]:
    s_lower = s.lower()
    key = []
    for ch in s_lower:
        if ch in TURKISH_ORDER:
            key.append(TURKISH_ORDER[ch])
        else:
            key.append(ord(ch) + 1000)
    return key


def russian_collation_key(s: str) -> List[int]:
    s_lower = s.lower()
    key = []
    for ch in s_lower:
        if ch in RUSSIAN_ORDER:
            key.append(RUSSIAN_ORDER[ch])
        else:
            key.append(ord(ch) + 1000)
    return key


def get_collation_key(word: str, language_pair: str) -> List[int]:
    source = language_pair.split("-")[0]
    if source == "ru":
        return russian_collation_key(word)
    elif source == "tr":
        return turkish_collation_key(word)
    else:
        return [ord(ch) for ch in word.lower()]


def compare_entries(e1: DictionaryEntry, e2: DictionaryEntry) -> int:
    key1 = get_collation_key(e1.word, e1.language_pair)
    key2 = get_collation_key(e2.word, e2.language_pair)
    for a, b in zip(key1, key2):
        if a < b:
            return -1
        elif a > b:
            return 1
    if len(key1) < len(key2):
        return -1
    elif len(key1) > len(key2):
        return 1
    else:
        return 0


# ------------------ Auto-Tagging ------------------

class AutoTagger:
    RUSSIAN_VERB_SUFFIXES = ("ать", "ять", "еть", "ить", "ыть", "уть", "оть", "ти", "чь")
    RUSSIAN_NOUN_SUFFIXES = ("а", "я", "ь", "ий", "ие", "ия", "о", "е", "м", "й", "ы", "и")
    RUSSIAN_ADJ_SUFFIXES = ("ый", "ий", "ой", "ая", "ее", "ие", "ые", "ое", "ское", "цкое")
    RUSSIAN_ADVERB_SUFFIXES = ("о", "е", "ски", "цки", "ому", "ему")
    RUSSIAN_PRONOUNS = ("я", "ты", "он", "она", "оно", "мы", "вы", "они", "мой", "твой", "его", "её", "наш", "ваш", "их")
    RUSSIAN_PREPOSITIONS = ("в", "на", "по", "за", "над", "под", "о", "об", "от", "до", "из", "без", "для", "к", "с", "у")
    RUSSIAN_CONJUNCTIONS = ("и", "а", "но", "или", "если", "что", "чтобы", "потому", "так", "как")
    RUSSIAN_INTERJECTIONS = ("ах", "ой", "ух", "эх", "ну", "вот", "эй", "ох")

    TURKISH_VERB_SUFFIXES = ("mak", "mek", "ar", "er", "ır", "ir", "ur", "ür", "t", "d", "yor", "di", "miş", "ecek")
    TURKISH_NOUN_SUFFIXES = ("lık", "lik", "luk", "lük", "ci", "cı", "cu", "cü", "çi", "çı", "çu", "çü", "daş", "deş")
    TURKISH_ADJ_SUFFIXES = ("li", "lı", "lu", "lü", "siz", "sız", "suz", "süz", "sel", "sal", "il", "ıl", "ul", "ül")
    TURKISH_ADVERB_SUFFIXES = ("ce", "ca", "çe", "ça", "en", "an", "ın", "in", "un", "ün", "ken")
    TURKISH_PRONOUNS = ("ben", "sen", "o", "biz", "siz", "onlar", "bu", "şu", "o")
    TURKISH_PREPOSITIONS = ("ile", "gibi", "kadar", "için", "üzere", "değin", "doğru", "karşı")
    TURKISH_CONJUNCTIONS = ("ve", "veya", "ama", "fakat", "çünkü", "ki", "ancak", "lâkin", "hem")
    TURKISH_INTERJECTIONS = ("hey", "oh", "ah", "vay", "ey", "ya", "of")

    @classmethod
    def tag(cls, word: str, language: str) -> str:
        word_lower = word.lower()
        if language == "ru":
            if any(word_lower.endswith(suffix) for suffix in cls.RUSSIAN_VERB_SUFFIXES):
                return "Verb"
            if any(word_lower.endswith(suffix) for suffix in cls.RUSSIAN_ADJ_SUFFIXES):
                return "Adjective"
            if any(word_lower.endswith(suffix) for suffix in cls.RUSSIAN_ADVERB_SUFFIXES):
                return "Adverb"
            if word_lower in cls.RUSSIAN_PRONOUNS:
                return "Pronoun"
            if word_lower in cls.RUSSIAN_PREPOSITIONS:
                return "Preposition"
            if word_lower in cls.RUSSIAN_CONJUNCTIONS:
                return "Conjunction"
            if word_lower in cls.RUSSIAN_INTERJECTIONS:
                return "Interjection"
            if any(word_lower.endswith(suffix) for suffix in cls.RUSSIAN_NOUN_SUFFIXES):
                return "Noun"
            return "Noun"
        elif language == "tr":
            if any(word_lower.endswith(suffix) for suffix in cls.TURKISH_VERB_SUFFIXES):
                return "Verb"
            if any(word_lower.endswith(suffix) for suffix in cls.TURKISH_NOUN_SUFFIXES):
                return "Noun"
            if any(word_lower.endswith(suffix) for suffix in cls.TURKISH_ADJ_SUFFIXES):
                return "Adjective"
            if any(word_lower.endswith(suffix) for suffix in cls.TURKISH_ADVERB_SUFFIXES):
                return "Adverb"
            if word_lower in cls.TURKISH_PRONOUNS:
                return "Pronoun"
            if word_lower in cls.TURKISH_PREPOSITIONS:
                return "Preposition"
            if word_lower in cls.TURKISH_CONJUNCTIONS:
                return "Conjunction"
            if word_lower in cls.TURKISH_INTERJECTIONS:
                return "Interjection"
            return "Noun"
        else:
            return "Noun"


# ------------------ Import/Export ------------------

class ImportExport:
    @staticmethod
    def export_json(entries: List[DictionaryEntry], filepath: str):
        data = [e.to_dict() for e in entries]
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    @staticmethod
    def import_json(filepath: str) -> List[DictionaryEntry]:
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                data = json.load(f)
        except json.JSONDecodeError as e:
            raise ValueError(f"JSON dosyası bozuk: {e}")
        except OSError as e:
            raise ValueError(f"Dosya okunamadı: {e}")
        if not isinstance(data, list):
            raise ValueError("JSON dosyası bir liste içermeli.")
        return [DictionaryEntry.from_dict(item) for item in data]

    @staticmethod
    def export_excel(entries: List[DictionaryEntry], filepath: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "Dictionary"
        ws.append(["Word", "Language Pair", "Status", "Is Favorite", "Part of Speech", "Definition", "Example"])
        for entry in entries:
            for trans in entry.translations:
                ws.append([
                    entry.word,
                    entry.language_pair,
                    entry.status,
                    "Yes" if entry.is_favorite else "No",
                    trans.part_of_speech,
                    trans.definition,
                    trans.example,
                ])
        wb.save(filepath)

    @staticmethod
    def import_excel(filepath: str) -> List[DictionaryEntry]:
        wb = load_workbook(filepath)
        ws = wb.active
        entries_dict: Dict[Tuple[str, str], List[Translation]] = {}
        status_dict: Dict[Tuple[str, str], str] = {}
        fav_dict: Dict[Tuple[str, str], bool] = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            if len(row) >= 7:
                word, pair, status, is_fav, pos, definition, example = row[:7]
            else:
                word, pair, status, pos, definition, example = row[:6]
                is_fav = False
            if not word or not pair:
                continue
            key = (word, pair)
            trans = Translation(pos or "", definition or "", example or "")
            if key not in entries_dict:
                entries_dict[key] = []
            entries_dict[key].append(trans)
            if status:
                status_dict[key] = status
            # FIX: Proper boolean check - string "No" is truthy in Python!
            if is_fav in (True, "Yes", "yes", "YES", "EVET", "evet", 1, "1"):
                fav_dict[key] = True

        entries = []
        for key, trans_list in entries_dict.items():
            st = status_dict.get(key, "not_started")
            is_fav = fav_dict.get(key, False)
            entries.append(DictionaryEntry(key[0], key[1], trans_list, st, is_fav))
        return entries

    @staticmethod
    def export_txt(entries: List[DictionaryEntry], filepath: str):
        with open(filepath, "w", encoding="utf-8") as f:
            for entry in entries:
                f.write(f"WORD: {entry.word}\n")
                f.write(f"PAIR: {entry.language_pair}\n")
                f.write(f"STATUS: {entry.status}\n")
                f.write(f"FAVORITE: {'Yes' if entry.is_favorite else 'No'}\n")
                for trans in entry.translations:
                    f.write(f"  [POS: {trans.part_of_speech}] {trans.definition}")
                    if trans.example:
                        f.write(f" (e.g., {trans.example})")
                    f.write("\n")
                f.write("\n")

    @staticmethod
    def import_txt(filepath: str) -> List[DictionaryEntry]:
        with open(filepath, "r", encoding="utf-8") as f:
            lines = f.readlines()
        entries = []
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            if not line:
                i += 1
                continue
            if line.startswith("WORD:"):
                word = line[len("WORD:"):].strip()
                i += 1
                while i < len(lines) and not lines[i].strip().startswith("PAIR:"):
                    i += 1
                if i >= len(lines):
                    break
                pair = lines[i].strip()[len("PAIR:"):].strip()
                i += 1
                status = "not_started"
                while i < len(lines) and not lines[i].strip().startswith("STATUS:"):
                    i += 1
                if i < len(lines) and lines[i].strip().startswith("STATUS:"):
                    status = lines[i].strip()[len("STATUS:"):].strip()
                    i += 1
                is_favorite = False
                while i < len(lines) and not lines[i].strip().startswith("FAVORITE:"):
                    if lines[i].strip().startswith("["):
                        break
                    i += 1
                if i < len(lines) and lines[i].strip().startswith("FAVORITE:"):
                    fav_val = lines[i].strip()[len("FAVORITE:"):].strip().lower()
                    is_favorite = fav_val in ("yes", "true", "1", "evet")
                    i += 1
                translations = []
                while i < len(lines):
                    line = lines[i].strip()
                    if line.startswith("WORD:") or line.startswith("PAIR:") or line.startswith("STATUS:"):
                        break
                    if not line:
                        i += 1
                        continue
                    m = re.match(
                        r"\s*\[POS:\s*([^\]]+)\]\s*(.*?)(?:\s*\(e\.g\.,\s*(.*)\))?$",
                        line,
                    )
                    if m:
                        pos = m.group(1).strip()
                        definition = m.group(2).strip()
                        example = m.group(3).strip() if m.group(3) else ""
                        translations.append(Translation(pos, definition, example))
                    i += 1
                if translations:
                    entries.append(DictionaryEntry(word, pair, translations, status, is_favorite))
            else:
                i += 1
        return entries
